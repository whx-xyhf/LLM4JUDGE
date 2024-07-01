import os
import openpyxl

from difflib import SequenceMatcher
def GetSimilarity(text1, text2):
    return SequenceMatcher(None, text1, text2).ratio()


def MatchParameter(LStr) :
    folder_path = 'data/TestTables'
    # 获取指定文件夹下的所有文件
    for filename in os.listdir(folder_path):
        # 检查文件是否是Excel文件
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # 拼接完整文件路径
            file_path = os.path.join(folder_path, filename)
        
            # #测试当前路径
            # print(file_path)

            file = openpyxl.load_workbook(file_path)   #加载.xlsx
            sheet = file.worksheets[0] #获取excel文件的第一个表单

            rows = sheet.max_row #表格中的最大行数
            syst = 2    #系统名称列
            name = 4    #故障现象列

            flag = 1

            for i,s in enumerate(LStr) :
                for j in range(rows):
                    varSys = sheet.cell(row=j+1, column=syst).value
                    varNa = sheet.cell(row=j+1, column=name).value
                    #匹配设备名字
                    if varSys is not None and GetSimilarity(LStr[i], varSys)>0.8: 
                        
                        #匹配现象名字
                        for u,s2 in enumerate(LStr) :
                            if varNa is not None and GetSimilarity(s2, varNa)>0.8:
                                
                                if flag == 1 :
                                    print(f"匹配现象是：{varNa}")
                                    flag-=1

                                param = ""
                                #运行参数
                                for col in range(6,14):
                                    if col < 13:

                                        if sheet.cell(row=1, column=col).value is not None :
                                            param = param + sheet.cell(row=1, column=col).value + '是'

                                        if sheet.cell(row=j+1, column=col).value is not None :
                                            param = param + str(sheet.cell(row=j+1, column=col).value ) + ';'
                                        else:
                                            param = param +'空' + '；'
                                            
                                    else:
                                        if sheet.cell(row=1, column=col).value is not None :
                                            param = param + sheet.cell(row=1, column=col).value + '是'

                                        if sheet.cell(row=j+1, column=col).value is not None :
                                            param = param + str(sheet.cell(row=j+1, column=col).value ) + ';'
                                        else:
                                            param = param +'空' + '。'
                                return param
                    
    return '空'

def MatchFactoryA(LStr):
    file_path = 'data/factoryRef/ft_factory_data.xlsx'
    # #测试当前路径
    # print(file_path)

    file = openpyxl.load_workbook(file_path)   #加载.xlsx
    sheet = file.worksheets[0] #获取excel文件的第一个表单
    rows = sheet.max_row #表格中的最大行数
    sys = 2         #系统(设备)名称列
    phnm = 8        #故障现象列
    condition = 14  #判据条件列
    for i,s in enumerate(LStr) :
        for j in range(rows):
            varSys = sheet.cell(row=j+1, column=sys).value
            varPh = sheet.cell(row=j+1, column=phnm).value
            #匹配设备名字
            if varSys is not None and GetSimilarity(LStr[i], varSys)>0.8: 
                #匹配现象名字
                for u,s2 in enumerate(LStr) :
                    if varPh is not None and GetSimilarity(s2, varPh)>0.8:
                        return sheet.cell(row=j+1, column=condition).value   
    return '空'

def MatchFactoryB(LStr):
    file_path = 'data/factoryRef/te_factory_data.xlsx'
    # #测试当前路径
    # print(file_path)

    file = openpyxl.load_workbook(file_path)   #加载.xlsx
    sheet = file.worksheets[0] #获取excel文件的第一个表单
    rows = sheet.max_row #表格中的最大行数
    sys = 2         #系统(设备)名称列
    phnm = 8        #故障现象列
    condition = 14  #判据条件列
    for i,s in enumerate(LStr) :
        for j in range(rows):
            varSys = sheet.cell(row=j+1, column=sys).value
            varPh = sheet.cell(row=j+1, column=phnm).value
            #匹配设备名字
            if varSys is not None and GetSimilarity(LStr[i], varSys)>0.8: 
                #匹配现象名字
                for u,s2 in enumerate(LStr) :
                    if varPh is not None and GetSimilarity(s2, varPh)>0.8:
                        return sheet.cell(row=j+1, column=condition).value   
    return '空'



#这个方法的作用是提取新工厂的故障现象的名称
def ExtractName(strs):
    chunk = strs
    singleStrs = chunk.split('\n')
    flag = 1
    for i, inStr in enumerate(singleStrs):
        child_str1 = "a)"
        #判断单条列表中是否含有 "a)" 子串 若有->停止判断，进入下一个列表
        if child_str1 in singleStrs[i] :
            # print(f"index: {index+1}___故障设备现象：{singleStrs[i-1]}")
            val_fault = singleStrs[i-1]
            pos = val_fault.find(' ')
            val_fault = val_fault[pos+1:]
            return val_fault

#这个方法的作用是提取新工厂的故障现象
def ExtractNewFactory_Fault_phnm(strs):
    alphabet = ['a','b','c','d','e']
    #分割出完成的一条数据 形成一个列表
    singleStrs = strs.split('\n')

    #列表数据测试
    # for i, inStr in enumerate(singleStrs):
    #     print(f"str[{i}] == {inStr}")
    # print()

    #第一个for：查询有原因现象的那一条数据
    for i, inStr in enumerate(singleStrs):
        flag = 'a'
        #如果 字符"a"出现在当前的字符串 inStr 中 则开始裁剪字符串
        if flag in inStr:
            faults = []
            #第二个for：便利所有出现的字母 截取原因 现象
            for j in range(len(alphabet)-1):
                c1 = alphabet[j]
                c2 = alphabet[j+1]
                index1 = 0
                index2 = 0
                if c1 in inStr:
                    index1 = inStr.index(c1)
                else:
                    break

                if c2 in inStr:
                    index2 = inStr.index(c2)
                    faults.append(inStr[index1:index2-1])
                else:
                    faults.append(inStr[index1:len(inStr)])
                    break

            #故障 现象 原因 内容测试
            # for k, str in enumerate(faults):
            #     print(str)
            # print()
            for k, str in enumerate(faults):
                c1 = "现象"
                s = str[0: 6]
                if c1 in s:
                    return str
            return '空'
           

#这个方法的作用是提取新工厂的故障原因
def ExtractNewFactory_Fault_reason(strs):
    alphabet = ['a','b','c','d','e']
    #分割出完成的一条数据 形成一个列表
    singleStrs = strs.split('\n')

    #列表数据测试
    # for i, inStr in enumerate(singleStrs):
    #     print(f"str[{i}] == {inStr}")
    # print()

    #第一个for：查询有原因现象的那一条数据
    for i, inStr in enumerate(singleStrs):
        flag = 'a'
        #如果 字符"a"出现在当前的字符串 inStr 中 则开始裁剪字符串
        if flag in inStr:
            faults = []
            #第二个for：便利所有出现的字母 截取原因 现象
            for j in range(len(alphabet)-1):
                c1 = alphabet[j]
                c2 = alphabet[j+1]
                index1 = 0
                index2 = 0
                if c1 in inStr:
                    index1 = inStr.index(c1)
                else:
                    break

                if c2 in inStr:
                    index2 = inStr.index(c2)
                    faults.append(inStr[index1:index2-1])
                else:
                    faults.append(inStr[index1:len(inStr)])
                    break

            #故障 现象 原因 内容测试
            # for k, str in enumerate(faults):
            #     print(str)
            # print()

            for k, str in enumerate(faults):
                # print(str)
                c2 = "原因"
                s = str[0: 6]
                if c2 in s:
                    # print(f"原因:____{str}")
                    return str
            return '空'

#下面这个算法的作用是将chunks2每一条故障名字、现象、原因 做个分割
def SplitData(chunk) :
    text = chunk.split('\n')
    for i, s in enumerate(text):
        if s != '' and s[0]!='a':
            pos = s.find(' ')
            text[i] = s[pos+1:]
    return text

        

        
    