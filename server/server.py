from flask import Flask, request, jsonify, make_response, render_template
# from flask_cors import CORS
# from flask_socketio import SocketIO, emit
# from flask_compress import Compress
from algorithm.chatGLM import get_glm_respones
import openpyxl
import json
import numpy as np

app = Flask(__name__)
# app.config['COMPRESS_ALGORITHM'] = 'gzip'
app.config['JSON_AS_ASCII'] = False
# CORS(app)
# compress = Compress()
# compress.init_app(app)
# socketio = SocketIO(app, cors_allowed_origins="*")

#目前先用这种变量存数据，所以只支持一个人用
global_data = {'guicheng':[],'dingzhiqingce':[], 'example':[]}

@app.after_request
def cors(environ):
    environ.headers['Access-Control-Allow-Origin'] = '*'
    environ.headers['Access-Control-Allow-Method'] = '*'
    environ.headers['Access-Control-Allow-Headers'] = 'x-requested-with,content-type'
    return environ


@app.route("/getResult", methods=['POST'])
def getResult():
    total = request.get_json()['total']
    limit = request.get_json()['limit']
    data, status = get_glm_respones(global_data, total=total, limit=limit)
    res = make_response(jsonify({'code': 200, 'data': data, 'status': status}))
    return res


#规程文件上传，暂时接收处理好的json文件
@app.route("/api/uploadGuicheng", methods=['POST'])
def uploadGuicheng():
    file = request.files.get('file')
    data = json.load(file)

    data = data.get("chunks", [])
    # print(chunks) #数据测试
    global_data['guicheng'].append(data)
    res = make_response(jsonify({'code': 200, 'data': ''}))
    return res

@app.route("/api/uploadExample", methods=['POST'])
def uploadExample():
    file = request.files.get('file')
    data_excel = openpyxl.load_workbook(file)  # 这里可以直接读文件对象
    data_sheet = data_excel[data_excel.sheetnames[0]]  # TODO 这是是读取第一个sheet的数据
    maxRows = data_sheet.max_row  # 行数
    maxColumns = data_sheet.max_column  # 列数
    # print(maxRows, maxColumns
    
    column_list = []
    for i in range(2, maxRows + 1):  # TODO 在这里不读取前面1行的str,只读取后面的数据
        for j in range(1, maxColumns + 1):
            column_list.append(data_sheet.cell(i, j).value)
    data = np.asarray(column_list).reshape(maxRows-1, maxColumns).astype(str) 
    global_data['example'].append(data) #读取数据存在global_data里
    
    #数据测试
    # para = global_data['dingzhiqingce']
    # print(para[0])
    # print("---------------------------")
    # for i,u in enumerate(para):
    #     for j,v in enumerate(u):
    #         for k,w in enumerate(v):
    #             print(f"para[{i}{j}{k}] = {w}")
    #             print()

    # print("---------------------------")
    
    res = make_response(jsonify({'code': 200, 'data': ''}))
    return res

#此处同上，都是读excel文件,只是存在global_data里的不同位置
###     dingzhiqingce
@app.route("/api/uploadValue", methods=['POST'])
def uploadValue():
    file = request.files.get('file')
    data_excel = openpyxl.load_workbook(file)  # 这里可以直接读文件对象
    data_sheet = data_excel[data_excel.sheetnames[0]]  # TODO 这是是读取第一个sheet的数据
    maxRows = data_sheet.max_row  # 行数
    maxColumns = data_sheet.max_column  # 列数
    # print(maxRows, maxColumns)

    column_list = []
    for i in range(1, maxRows + 1):  # TODO 
        for j in range(1, maxColumns + 1):
            column_list.append(data_sheet.cell(i, j).value)
    
    data = np.asarray(column_list).reshape(maxRows, maxColumns).astype(str) 
    
    name = file.filename
    new_dict = {name: data}
    # print(new_dict)

    global_data['dingzhiqingce'].append(new_dict) #读取数据存在global_data里
    # print(global_data)
    # para = global_data['dingzhiqingce']
    # print(para[0])

    res = make_response(jsonify({'code': 200, 'data': ''}))

    return res

# @socketio.on('send_message')
# def handle_message(message):  # 其中的message是前端传过来的
#     print('Received message: ' + message)  # 打印前端传来的信息
#     response = 'Server received: ' + message  # 字符串拼接
#     emit('message', response)


@app.route('/', methods=['get'])
def root():
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True, host="0.0.0.0", port=5000)
    # socketio.run(app, debug=True, allow_unsafe_werkzeug=True, host="0.0.0.0", port=5000)

